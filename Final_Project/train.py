import torch
import tqdm
from config import args
from sklearn.metrics import roc_auc_score
import csv
import ast
import torch.nn as nn
import torch.optim as optim
import openpyxl
import os
import numpy as np
import pandas as pd
from sklearn.metrics import roc_auc_score, accuracy_score
from utils import merge_csv_columns


num_users, num_items = args.user_length, args.item_length
epochs = 10

class EarlyStopper(object):

    def __init__(self, num_trials, save_path):
        self.num_trials = num_trials
        self.trial_counter = 0
        self.best_accuracy = 0
        self.save_path = save_path

    def is_continuable(self, model, accuracy):
        if accuracy > self.best_accuracy:
            self.best_accuracy = accuracy
            self.trial_counter = 0
            torch.save(model, self.save_path) # best model
            return True
        elif self.trial_counter + 1 < self.num_trials:
            self.trial_counter += 1
            return True
        else:
            return False
# Define the model architecture
class FullyConnectedModel(nn.Module):
    def __init__(self, batch_size, output_dim=1):
        super(FullyConnectedModel, self).__init__()
        self.fc = nn.Linear(batch_size, output_dim, bias=False)
        self.user_bias = nn.Parameter(torch.zeros(num_users))
        self.item_bias = nn.Parameter(torch.zeros(num_items))
        self.global_bias = nn.Parameter(torch.zeros(1))
        # self.sigmoid = nn.Sigmoid()

    def forward(self, user_indices, item_indices, user_features, item_features):
        # Element-wise multiplication of user and item features
        user_features = torch.tensor(user_features)
        # print(user_features.shape)
        item_features = torch.tensor(item_features)
        interaction = user_features * item_features
        # Sum over features
        
        interaction = interaction.sum(dim=1)
        # print(interaction.shape)
        self.fc = nn.Linear(len(interaction), len(interaction), bias=True)
        # Multiply by weights
        prediction = self.fc(interaction.to(dtype=torch.float32))
        # Add biases
        prediction += self.global_bias + self.user_bias[user_indices] + self.item_bias[item_indices]
        # prediction = self.sigmoid(prediction.squeeze(0))
        
        return prediction.squeeze()

def save_to_excel(values, headers, filename):
    # Kiểm tra xem tệp đã tồn tại hay không
    if os.path.exists(filename):
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active
        start_row = sheet.max_row  # Bắt đầu từ hàng tiếp theo
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        start_row = 1  # Bắt đầu từ hàng đầu tiên

        # Ghi header cho từng cột nếu tệp chưa tồn tại
        for index, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=index, value=header)

    # Append các giá trị (đã làm tròn) vào các hàng tiếp theo
    for row_index, row_values in enumerate(values, start=start_row + 1):
        for col_index, value in enumerate(row_values, start=1):
            sheet.cell(row=row_index, column=col_index, value=round(value, 4))

    # Lưu workbook vào tệp Excel
    workbook.save(filename)
    print(f"Đã lưu danh sách giá trị vào tệp '{filename}' thành công.")

def reprocess_input(data, isTrain=True):
    user_idx = [int(x) for x in data['reviewerID']]
    item_idx = [int(x) for x in data['itemID']]
    rating = [float(x) for x in data['overall']]
    if isTrain:
        list_udeep = []
        for item in data['Udeep']:
            list_udeep.append(np.array(ast.literal_eval(item)))
        user_feature = list_udeep
        list_ideep = []
        for item in data['Ideep']:
            list_ideep.append(np.array(ast.literal_eval(item)))
        item_feature = list_ideep
        return user_idx, item_idx, rating, user_feature, item_feature
    else:
        return user_idx, item_idx, rating

def load_data(csv_file):
    data = []
    with open(csv_file, 'r', newline='') as file:
        
        csv_reader = csv.DictReader(file)
        for row in csv_reader:
            data.append(row)
    return data

def calculate_rmse(y_true, y_pred):
    """
    Tính toán Root Mean Square Error (RMSE) giữa y_true và y_pred.
    
    Tham số:
    - y_true: mảng numpy chứa các giá trị thực tế.
    - y_pred: mảng numpy chứa các giá trị dự đoán.
    
    Trả về:
    - Giá trị RMSE.
    """
    # Chuyển đổi danh sách thành mảng NumPy
    y_true_np = np.array(y_true)
    y_pred_np = np.array(y_pred)

    # Tính toán squared_errors
    squared_errors = (y_true_np - y_pred_np) ** 2
    # Tính trung bình của bình phương sai số
    mean_squared_error = np.mean(squared_errors)
    # Tính RMSE bằng cách lấy căn bậc hai của trung bình bình phương sai số
    rmse = np.sqrt(mean_squared_error)
    return rmse

def test_rsme(model, data_loader, device):
    model.eval()
    targets, predicts = list(), list()
    with torch.no_grad():
        for data in tqdm.tqdm(data_loader, smoothing=0, mininterval=1.0):
            user_idx, item_idx, target, udeep, ideep = reprocess_input(data)
            y = model(user_idx, item_idx, udeep, ideep)
            targets.extend(target)
            # print("Max: ", max(y))
            # print("Min: ", min(y))
            predicts.extend([float(pred) for pred in y.flatten().cpu().numpy()])

    new_targer = []
    new_predict = []
    for i in targets:
        if i < 3:
            new_targer.append(-1)
        else:
            new_targer.append(1)
    for i in predicts:
        if i < 3:
            new_predict.append(-1)
        else:
            new_predict.append(1)
            
    return roc_auc_score(new_targer, new_predict) 
    # return roc_auc_score(targets, predicts) 
    # return calculate_rmse(targets, predicts)

def test(model, data_loader, device):
    model.eval()
    targets, predicts = list(), list()
    with torch.no_grad():
        for data in tqdm.tqdm(data_loader, smoothing=0, mininterval=1.0):
            user_idx, item_idx, target, udeep, ideep = reprocess_input(data)
            y = model(user_idx, item_idx, udeep, ideep )
            targets.extend(target)
            predicts.extend([round(float(pred)) for pred in y.flatten().cpu().numpy()])
    # print(targets)
    # print(predicts)
    new_targer = []
    new_predict = []
    for i in targets:
        if i < 3:
            new_targer.append(-1)
        else:
            new_targer.append(1)
    for i in predicts:
        if i < 3:
            new_predict.append(-1)
        else:
            new_predict.append(1)
            
    accuracy = accuracy_score(new_targer, new_predict)
    return accuracy

def train(model, optimizer, data_loader, criterion, device, log_interval=100):
    # Train the model
    model.train()
    total_loss = 0

    for batch in data_loader:
        try:
        # Xử lý batch ở đây
            user_idx, item_idx, rating, user_feature, item_feature = reprocess_input(batch)
            
            # Forward pass
            predictions = model(user_idx, item_idx, user_feature, item_feature)

            # Calculate loss
            loss = criterion(predictions, torch.tensor(rating))  # rating needs to be a column vector
            
            # Backward pass and optimize
            optimizer.zero_grad()
            loss.backward()
            optimizer.step()
            total_loss += loss.item()
        except Exception as e:
            # Code block to handle the exception if it occurs
            # This block will be executed if an exception of type ExceptionType is raised
            print("error: ",e)
            
    
    print(f"Epoch {epochs+1}, Loss: {total_loss / len(data_loader)}")


#---------------------------------Main---------------------------------
from deep_fusion import *
from torch.utils.data import DataLoader

# train_data = pd.read_csv(args.data_feature, sep=',', engine='c', header='infer', chunksize=args.batch_size)
train_data = load_data(args.data_feature)
valid_data = load_data('./feature/valid_data_df.csv')
test_data = load_data('./feature/test_data_df.csv')
# Tạo list chứa các DataFrame từ train_data_loader, valid_data_loader, test_data_loader
# train_dfs = [df for df in train_data_loader]
valid_dfs = [df for df in valid_data]
test_dfs = [df for df in test_data]

# Gán các cột 'Udeep' và 'Ideep' với giá trị mảng 0 có 10 phần tử
for df in [valid_dfs, test_dfs]:
    for df_chunk in df:
        df_chunk['Udeep'] = [([0] * 10) for _ in range(len(df_chunk))]
        df_chunk['Ideep'] = [([0] * 10) for _ in range(len(df_chunk))]
print(valid_dfs)
train_data_loader = DataLoader(train_data, batch_size=args.batch_size, num_workers=8)
valid_data_loader = DataLoader(valid_dfs, batch_size=args.batch_size, num_workers=8)
test_data_loader = DataLoader(test_dfs, batch_size=args.batch_size, num_workers=8)

model = FullyConnectedModel(args.batch_size, output_dim=20)
criterion = nn.MSELoss()
optimizer = optim.Adam(model.parameters(), lr=0.01)
filename = 'triet_deepcgsr'
early_stopper = EarlyStopper(num_trials=3, save_path=f'{args.save_dir}/{filename}.pt')
rsme = 0
count = 0
for i in range(10):
    for epoch_i in range(epochs):
        train(model, optimizer, train_data_loader, criterion, args.device)
        auc = test(model, valid_data_loader, args.device)
        print('epoch:', epoch_i, 'validation: auc:', auc)
        rsme = rsme + test_rsme(model, valid_data_loader, args.device)
        print('epoch:', epoch_i, 'validation: rsme:', rsme)
        count = count + 1
        if not early_stopper.is_continuable(model, auc):
            print(f'validation: best auc: {early_stopper.best_accuracy}')
            
            break

    print('epoch:', i, 'validation: rsme:', rsme/count)
    auc_test = test(model, test_data_loader, args.device)
    print(f'test auc: {auc_test}')
    rsme_test = test_rsme(model, test_data_loader, args.device)
    print(f'test rsme:', {rsme_test})
    print(f'average validate rsme:', {rsme/count})
    results = [auc_test, rsme_test]
    if args.isRemoveOutliner:
        save_to_excel([results], ['AUC', 'RSME Test'], 'results_outliner.xlsx')
    else:
        save_to_excel([results], ['AUC', 'RSME Test'], 'results_DigitalMusic.xlsx')
